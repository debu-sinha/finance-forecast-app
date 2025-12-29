"""
Excel Exporter for Simple Mode.

Export forecast to Excel with formulas - familiar format for finance users.
They can see the logic, make adjustments, share with stakeholders.
"""

import io
from typing import Dict, Any, List, Optional
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import pandas as pd

from .forecast_explainer import ForecastExplanation

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Export forecast to Excel with formulas.
    Familiar format for finance users who are used to Excel.
    """

    # Style constants
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def export(
        self,
        forecast_result: Dict[str, Any],
        explanation: ForecastExplanation,
        input_data: Optional[pd.DataFrame] = None,
        slice_forecasts: Optional[List[Dict[str, Any]]] = None
    ) -> bytes:
        """
        Generate Excel file with multiple sheets.

        Args:
            forecast_result: Raw forecast results
            explanation: ForecastExplanation object
            input_data: Optional original input data
            slice_forecasts: Optional list of slice forecast results (for by-slice mode)

        Returns:
            Excel file as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        logger.info("Generating Excel export...")

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Sheet 1: Summary (executive view)
        self._add_summary_sheet(wb, forecast_result, explanation, slice_forecasts)

        # Sheet 2: Forecast Detail (period-by-period with formulas)
        self._add_forecast_detail_sheet(wb, explanation)

        # Sheet 3: Component Breakdown
        self._add_decomposition_sheet(wb, explanation)

        # Sheet 4: Confidence & Quality
        self._add_confidence_sheet(wb, explanation)

        # Sheet 5: Audit Trail
        self._add_audit_sheet(wb, explanation.audit_trail)

        # Sheet 6: Raw Data (if provided)
        if input_data is not None:
            self._add_raw_data_sheet(wb, input_data)

        # Add individual slice forecast sheets (if by-slice mode)
        if slice_forecasts and len(slice_forecasts) > 0:
            logger.info(f"Adding {len(slice_forecasts)} slice forecast sheets...")
            for slice_data in slice_forecasts:
                self._add_slice_forecast_sheet(wb, slice_data)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel export complete: {len(wb.sheetnames)} sheets")

        return output.getvalue()

    def _add_summary_sheet(
        self, wb: Workbook, result: Dict[str, Any], explanation: ForecastExplanation,
        slice_forecasts: Optional[List[Dict[str, Any]]] = None
    ):
        """Create executive summary sheet."""

        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = "FORECAST SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Generated timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, color="666666")

        # Key metrics section
        row = 4
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.SUBHEADER_FILL

        row += 1

        # Check if this is a by-slice forecast
        is_by_slice = slice_forecasts and len(slice_forecasts) > 0

        metrics = [
            ("Total Forecast", f"${sum(result.get('forecast', [])):,.0f}"),
            ("Forecast Horizon", f"{len(result.get('forecast', []))} periods"),
            ("Best Model", result.get('best_model', 'N/A')),
            ("Accuracy (MAPE)", f"{result.get('metrics', {}).get('mape', 'N/A'):.1f}%"),
            ("Confidence Level", explanation.confidence.level.capitalize()),
            ("Confidence Score", f"{explanation.confidence.score:.0f}/100"),
        ]

        # Add slice count for by-slice mode
        if is_by_slice:
            metrics.insert(2, ("Forecast Mode", f"By-Slice ({len(slice_forecasts)} segments)"))

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Add slice summary section if by-slice mode
        if is_by_slice:
            row += 1
            ws[f'A{row}'] = "SLICE BREAKDOWN"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = self.SUBHEADER_FILL

            row += 1
            # Headers
            slice_headers = ["Slice", "Model", "MAPE", "Avg Forecast", "Data Points"]
            for col, header in enumerate(slice_headers):
                ws.cell(row=row, column=col+1, value=header)
                ws.cell(row=row, column=col+1).fill = self.HEADER_FILL
                ws.cell(row=row, column=col+1).font = self.HEADER_FONT

            row += 1
            for slice_data in slice_forecasts:
                ws.cell(row=row, column=1, value=slice_data.get('slice_id', 'Unknown'))
                ws.cell(row=row, column=2, value=slice_data.get('best_model', 'N/A'))
                mape = slice_data.get('holdout_mape')
                ws.cell(row=row, column=3, value=f"{mape:.1f}%" if mape else 'N/A')
                forecast = slice_data.get('forecast', [])
                avg_forecast = sum(forecast) / len(forecast) if forecast else 0
                ws.cell(row=row, column=4, value=f"${avg_forecast:,.0f}")
                ws.cell(row=row, column=5, value=slice_data.get('data_points', 0))
                row += 1

            row += 1
            ws[f'A{row}'] = "Note: Each slice has its own forecast sheet below."
            ws[f'A{row}'].font = Font(italic=True, color="666666")

        # Caveats section
        row += 1
        ws[f'A{row}'] = "IMPORTANT NOTES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.SUBHEADER_FILL

        row += 1
        for caveat in explanation.caveats:
            ws[f'A{row}'] = f"• {caveat}"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

        # Reproducibility info
        row += 1
        ws[f'A{row}'] = "REPRODUCIBILITY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.SUBHEADER_FILL

        row += 1
        ws[f'A{row}'] = "Token"
        ws[f'B{row}'] = explanation.audit_trail.reproducibility_token
        ws[f'B{row}'].font = Font(name='Consolas', size=9)

        row += 1
        ws[f'A{row}'] = "Note"
        ws[f'B{row}'] = "Same token guarantees identical output when re-run"

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _add_forecast_detail_sheet(self, wb: Workbook, explanation: ForecastExplanation):
        """
        Create detailed forecast sheet with formulas.
        This is the key sheet that makes it feel like Excel.
        """

        ws = wb.create_sheet("Forecast Detail")

        # Headers
        headers = [
            "Date", "Forecast", "Lower Bound", "Upper Bound",
            "Base", "Trend", "Seasonal", "Holiday", "Formula Check"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER

        # Data rows with Excel formulas
        for i, period in enumerate(explanation.components.period_breakdown, 2):
            # Date
            ws.cell(row=i, column=1, value=str(period.period) if period.period else "")

            # Forecast value
            ws.cell(row=i, column=2, value=period.forecast)

            # Bounds
            ws.cell(row=i, column=3, value=period.lower_bound)
            ws.cell(row=i, column=4, value=period.upper_bound)

            # Components
            ws.cell(row=i, column=5, value=period.base)      # Base
            ws.cell(row=i, column=6, value=period.trend)     # Trend
            ws.cell(row=i, column=7, value=period.seasonal)  # Seasonal
            ws.cell(row=i, column=8, value=period.holiday)   # Holiday

            # Formula that shows how forecast is calculated
            # =E{row}+F{row}+G{row}+H{row}
            ws.cell(row=i, column=9, value=f"=E{i}+F{i}+G{i}+H{i}")

            # Apply borders
            for col in range(1, 10):
                ws.cell(row=i, column=col).border = self.BORDER

        # Format numbers
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=9):
            for cell in row:
                cell.number_format = '#,##0.00'

        # Add formula explanation row
        last_row = len(explanation.components.period_breakdown) + 3
        ws.cell(row=last_row, column=1, value="Formula:")
        ws.cell(row=last_row, column=1).font = Font(bold=True)
        ws.merge_cells(f'B{last_row}:I{last_row}')
        ws.cell(row=last_row, column=2, value=explanation.components.formula)

        # Column widths
        widths = [12, 12, 12, 12, 12, 10, 10, 10, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

    def _add_decomposition_sheet(self, wb: Workbook, explanation: ForecastExplanation):
        """Create component breakdown sheet with pie/bar visualization."""

        ws = wb.create_sheet("Components")

        # Title
        ws['A1'] = "FORECAST COMPONENT BREAKDOWN"
        ws['A1'].font = Font(size=14, bold=True)

        # Totals table
        ws['A3'] = "Component"
        ws['B3'] = "Total Value"
        ws['C3'] = "% of Total"
        for col in range(1, 4):
            ws.cell(row=3, column=col).fill = self.HEADER_FILL
            ws.cell(row=3, column=col).font = self.HEADER_FONT

        components = [
            ("Base", explanation.components.total_base),
            ("Trend", explanation.components.total_trend),
            ("Seasonal", explanation.components.total_seasonal),
            ("Holiday", explanation.components.total_holiday),
        ]

        total = sum(abs(c[1]) for c in components)

        for i, (name, value) in enumerate(components, 4):
            ws.cell(row=i, column=1, value=name)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=2).number_format = '#,##0.00'
            pct = (abs(value) / total * 100) if total > 0 else 0
            ws.cell(row=i, column=3, value=f"{pct:.1f}%")

        # Add total row
        total_row = len(components) + 4
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f"=SUM(B4:B{total_row-1})")
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=2).number_format = '#,##0.00'

        # Explanation
        ws['A10'] = "How to read this:"
        ws['A10'].font = Font(bold=True)
        ws['A11'] = "• Base: The average expected value based on historical data"
        ws['A12'] = "• Trend: The growth or decline component"
        ws['A13'] = "• Seasonal: Regular patterns (weekly, monthly, yearly)"
        ws['A14'] = "• Holiday: Special holiday effects (Thanksgiving, Christmas, etc.)"

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12

    def _add_confidence_sheet(self, wb: Workbook, explanation: ForecastExplanation):
        """Create confidence and quality assessment sheet."""

        ws = wb.create_sheet("Confidence")

        # Title
        ws['A1'] = "CONFIDENCE ASSESSMENT"
        ws['A1'].font = Font(size=14, bold=True)

        # Overall confidence
        ws['A3'] = "Overall Confidence"
        ws['B3'] = explanation.confidence.level.upper()
        ws['A3'].font = Font(bold=True)

        # Color code confidence
        if explanation.confidence.level == "high":
            ws['B3'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        elif explanation.confidence.level == "medium":
            ws['B3'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        else:
            ws['B3'].fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

        ws['A4'] = "Confidence Score"
        ws['B4'] = f"{explanation.confidence.score:.0f}/100"

        ws['A5'] = "Model Accuracy (MAPE)"
        ws['B5'] = f"{explanation.confidence.mape:.1f}%"

        # Factors table
        ws['A7'] = "CONFIDENCE FACTORS"
        ws['A7'].font = Font(bold=True, size=12)
        ws['A7'].fill = self.SUBHEADER_FILL

        headers = ["Factor", "Score", "Note"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        for i, factor in enumerate(explanation.confidence.factors, 9):
            ws.cell(row=i, column=1, value=factor['factor'])
            ws.cell(row=i, column=2, value=f"{factor['score']}/100")
            ws.cell(row=i, column=3, value=factor['note'])

        # Explanation
        last_row = 9 + len(explanation.confidence.factors) + 1
        ws.cell(row=last_row, column=1, value="Explanation:")
        ws.cell(row=last_row, column=1).font = Font(bold=True)
        ws.merge_cells(f'A{last_row+1}:C{last_row+1}')
        ws.cell(row=last_row+1, column=1, value=explanation.confidence.explanation)

        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30

    def _add_audit_sheet(self, wb: Workbook, audit_trail):
        """Create audit trail sheet for compliance."""

        ws = wb.create_sheet("Audit Trail")

        # Title
        ws['A1'] = "AUDIT TRAIL"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A2'] = "For compliance and reproducibility"
        ws['A2'].font = Font(italic=True, color="666666")

        # Audit details
        audit_items = [
            ("Run ID", audit_trail.run_id),
            ("Timestamp", audit_trail.run_timestamp),
            ("", ""),  # Spacer
            ("INPUT DATA", ""),
            ("Data Hash", audit_trail.input_data_hash),
            ("Row Count", str(audit_trail.input_row_count)),
            ("Date Range", f"{audit_trail.input_date_range[0]} to {audit_trail.input_date_range[1]}"),
            ("", ""),  # Spacer
            ("CONFIGURATION", ""),
            ("Config Hash", audit_trail.config_hash),
            ("", ""),  # Spacer
            ("MODEL", ""),
            ("Model Type", audit_trail.model_type),
            ("Model Version", audit_trail.model_version),
            ("MLflow Run ID", audit_trail.mlflow_run_id or "N/A"),
            ("Model URI", audit_trail.model_uri or "N/A"),
            ("", ""),  # Spacer
            ("OUTPUT", ""),
            ("Output Hash", audit_trail.output_hash),
            ("", ""),  # Spacer
            ("REPRODUCIBILITY", ""),
            ("Token", audit_trail.reproducibility_token),
        ]

        for i, (label, value) in enumerate(audit_items, 4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)

            # Style section headers
            if value == "" and label != "":
                ws.cell(row=i, column=1).font = Font(bold=True, size=11)
                ws.cell(row=i, column=1).fill = self.SUBHEADER_FILL
            elif label != "":
                ws.cell(row=i, column=1).font = Font(bold=True)

            # Monospace font for hashes
            if "Hash" in label or "Token" in label or "ID" in label:
                ws.cell(row=i, column=2).font = Font(name='Consolas', size=9)

        # Reproducibility note
        last_row = 4 + len(audit_items) + 1
        ws.merge_cells(f'A{last_row}:B{last_row}')
        ws.cell(row=last_row, column=1, value="Note: Use the reproducibility token to regenerate this exact forecast.")
        ws.cell(row=last_row, column=1).font = Font(italic=True)

        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50

    def _add_raw_data_sheet(self, wb: Workbook, data: pd.DataFrame):
        """Add raw input data sheet."""

        ws = wb.create_sheet("Raw Data")

        # Title
        ws['A1'] = "INPUT DATA"
        ws['A1'].font = Font(size=14, bold=True)

        ws['A2'] = f"Total rows: {len(data)}"
        ws['A2'].font = Font(italic=True)

        # Add data starting from row 4
        for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 4):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Style header
                if r_idx == 4:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT

        # Auto-width columns (with max limit)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)

    def _add_slice_forecast_sheet(self, wb: Workbook, slice_data: Dict[str, Any]):
        """
        Add individual slice forecast sheet.

        Each slice gets its own sheet with:
        - Slice metadata (filters, model, accuracy)
        - Period-by-period forecast with confidence intervals
        - Summary statistics
        """
        slice_id = slice_data.get('slice_id', 'Unknown')

        # Create safe sheet name (Excel limits to 31 chars, no special chars)
        safe_name = slice_id[:28].replace('/', '-').replace('\\', '-').replace(':', '-')
        safe_name = safe_name.replace('[', '(').replace(']', ')').replace('*', '')
        safe_name = safe_name.replace('?', '').replace("'", "")

        # Handle duplicate sheet names by adding suffix
        base_name = safe_name
        counter = 1
        while safe_name in wb.sheetnames:
            safe_name = f"{base_name[:25]}_{counter}"
            counter += 1

        ws = wb.create_sheet(f"Slice - {safe_name}")

        # Title
        ws['A1'] = f"SLICE FORECAST: {slice_id}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')

        # Metadata section
        row = 3
        ws[f'A{row}'] = "SLICE DETAILS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.SUBHEADER_FILL

        row += 1

        # Slice filters
        filters = slice_data.get('slice_filters', {})
        if filters:
            ws[f'A{row}'] = "Filters"
            filter_str = " | ".join([f"{k}={v}" for k, v in filters.items()])
            ws[f'B{row}'] = filter_str
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Model info
        ws[f'A{row}'] = "Best Model"
        ws[f'B{row}'] = slice_data.get('best_model', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Accuracy
        mape = slice_data.get('holdout_mape')
        ws[f'A{row}'] = "Holdout MAPE"
        ws[f'B{row}'] = f"{mape:.2f}%" if mape is not None else 'N/A'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Data points
        ws[f'A{row}'] = "Training Data Points"
        ws[f'B{row}'] = slice_data.get('data_points', 0)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # Summary stats
        forecast = slice_data.get('forecast', [])
        if forecast:
            ws[f'A{row}'] = "Total Forecast"
            ws[f'B{row}'] = f"${sum(forecast):,.2f}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            ws[f'A{row}'] = "Avg Period Forecast"
            ws[f'B{row}'] = f"${sum(forecast)/len(forecast):,.2f}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

        # Forecast data section
        row += 1
        ws[f'A{row}'] = "FORECAST DATA"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.SUBHEADER_FILL

        row += 1

        # Headers
        headers = ["Date", "Forecast", "Lower Bound", "Upper Bound", "Range Width"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER

        row += 1
        data_start_row = row

        # Forecast data
        dates = slice_data.get('dates', [])
        lower_bounds = slice_data.get('lower_bounds', [])
        upper_bounds = slice_data.get('upper_bounds', [])

        for i, date in enumerate(dates):
            ws.cell(row=row, column=1, value=str(date))
            ws.cell(row=row, column=1).border = self.BORDER

            # Forecast
            forecast_val = forecast[i] if i < len(forecast) else 0
            ws.cell(row=row, column=2, value=forecast_val)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=2).border = self.BORDER

            # Lower bound
            lower = lower_bounds[i] if i < len(lower_bounds) else forecast_val
            ws.cell(row=row, column=3, value=lower)
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).border = self.BORDER

            # Upper bound
            upper = upper_bounds[i] if i < len(upper_bounds) else forecast_val
            ws.cell(row=row, column=4, value=upper)
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=4).border = self.BORDER

            # Range width formula (shows uncertainty)
            ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=5).border = self.BORDER

            row += 1

        # Add totals row
        if len(dates) > 0:
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).border = self.BORDER

            for col in range(2, 6):
                ws.cell(row=row, column=col, value=f"=SUM({chr(64+col)}{data_start_row}:{chr(64+col)}{row-1})")
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).number_format = '#,##0.00'
                ws.cell(row=row, column=col).border = self.BORDER

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15


def export_forecast_to_excel(
    forecast_result: Dict[str, Any],
    explanation: ForecastExplanation,
    input_data: Optional[pd.DataFrame] = None,
    slice_forecasts: Optional[List[Dict[str, Any]]] = None
) -> bytes:
    """
    Convenience function to export forecast to Excel.

    Args:
        forecast_result: Raw forecast results
        explanation: ForecastExplanation object
        input_data: Optional original input data
        slice_forecasts: Optional list of slice forecast results (for by-slice mode)

    Returns:
        Excel file as bytes
    """
    exporter = ExcelExporter()
    return exporter.export(forecast_result, explanation, input_data, slice_forecasts)
